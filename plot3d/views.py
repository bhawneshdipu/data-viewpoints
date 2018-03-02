from django.shortcuts import render,render_to_response,redirect

from django.template import RequestContext
from django.core.files.storage import FileSystemStorage

import  os
import openpyxl
from openpyxl import  Workbook
from openpyxl.utils import get_column_letter
from origin import settings
from django.core.files import File
from matplotlib import pyplot as plt
import numpy as np
from scipy import linspace, polyval, polyfit, sqrt, stats
from scipy.optimize import curve_fit
import matplotlib.mlab as mlab
from scipy.interpolate import  interp1d
# Create your views here.
from django.http import HttpResponse,HttpResponseRedirect,HttpRequest
import  pandas as pd
from django.views.decorators.csrf import csrf_exempt
import shutil
from bokeh.plotting import figure, output_file, show
from bokeh.resources import CDN
from bokeh.embed import components
from bokeh.embed import autoload_static
from bokeh.charts import Bar,BoxPlot,Area,Histogram,Line,Scatter,TimeSeries,Donut
import datetime,time
import glob
from functools import  reduce
from django.utils.html import escape
from itertools import cycle
from sklearn.cluster import AffinityPropagation,KMeans
from sklearn import metrics
from sklearn.datasets.samples_generator import make_blobs
from matplotlib.mlab import griddata
from mpl_toolkits.mplot3d import Axes3D



@csrf_exempt
def save_data(request):
    print("save_data")
    if request.method=='POST':
        rows=request.POST.get('rows')
        cols=request.POST.get('cols')
        #print(rows,cols)
        rows=int(rows)
        cols=int(cols)
        index_table(request,request.POST.get('data'),rows,cols)
        return HttpResponse(1)
    return HttpResponse(0)

def myfloat(seq,row,col):
    print('myfloat')
    ans=[]

    temp = []
    i=0
    j=0
    print(row,col)
    for x in seq:
        try:
            if col>j:
                if len(x)>0:
                    temp.append((x))
                    j+=1
                    print(x,j)
            else:
                j=1
                print('else')
                print(ans)
                print(x,j)
                ans.append(temp)
                print(ans)
                temp=[]
                if len(x)>0:
                    temp.append((x))
        except ValueError:
            print('value error')
            print(len(temp))
            print(temp)
            if len(temp)>0:
                ans.append(temp)
                print(temp)
                print(ans)
            j=0
            temp=[]
    if len(temp)>0:
        ans.append(temp)
    print(ans)
    return ans

def index_table(request,ddata,row,col):
    print('index table')
    rows=int(row)
    cols=int(col)

    username=request.COOKIES.get('user_name')
    logintime=request.COOKIES.get('login_time')
    print(username)
    print(logintime)
    dest_filename = 'plot3d/static/xlsx/'+username+logintime+'.xlsx'
    fs = FileSystemStorage(location='plot3d/static/xlsx/')
    if (fs.exists(username+logintime+'.xlsx')):
        os.remove('plot3d/static/xlsx/'+username+logintime+'.xlsx')

    print(ddata)
    data=ddata.split(",")
    print(data)
    data=myfloat(data,rows,cols)
    print(data)
    #data=list(data)
    #print(data)
    print('data')
    k=0
    l=0
    print(cols,rows)
    wb = Workbook()
    ws = wb.active
    print(len(data))
    print(len(data[0]))
    try:
        for i in range(0,len(data)):

                for j in range(0,len(data[i])+1):
                    try:
                        ws.cell(column=j+1, row=i+1,value=data[i][j])
                    except:
                        print('break')
                        break

        wb.save(dest_filename)
        print("saved")
    except:
        wb.save(dest_filename)
        print('saved with exception')


@csrf_exempt
def del_user(request):
    request.session.set_test_cookie()
    ddata = 0
    tabled = 0
    plotplot3d = 0
    username='not a valid user'
    logintime=0
    if 'user_name' in request.COOKIES:
        username=request.COOKIES.get('user_name')
    if 'login_time' in request.COOKIES:
        logintime=request.COOKIES.get('login_time')

    pathi='plot3d/static/img/'+username+'*.*'
    pathf = 'plot3d/static/xlsx/' + username + str(logintime) + '*.*'

    for hgx in glob.glob(pathi):
        os.remove(hgx)
    for hgx in glob.glob(pathf):
        os.remove(hgx)

    if request.session.test_cookie_worked():
        request.session.delete_test_cookie()
        response = HttpResponse("Bye")
        response.delete_cookie('user_name')
        response.delete_cookie('login_time')
        return  response
    else:
        return HttpResponse("Sorry")

@csrf_exempt
def default(request):
    if 'user_name' in request.COOKIES:
        username=request.COOKIES['user_name']
        return redirect('./plot3d','module.html',locals())
    elif request.method=='POST':
        username=request.POST.get('username')
        response=redirect('./plot3d','3d_index.html',request)
        response.set_cookie('user_name',request.POST.get('username'))
        response.set_cookie('login_time',time.time())
        #return redirect('./plot3d',request,locals())
        return response
    else:
        return render(request,'default.html',locals())



@csrf_exempt
def index(request):
    errormsg=''
    if request.method!='POST':
        print('not post')
        ddata=0
        tabled=0
        plotplot3d=0
        if 'user_name' in request.COOKIES and 'login_time' in request.COOKIES:
            username = request.COOKIES['user_name']
            logintime=request.COOKIES.get('login_time')
            dest_filename = 'plot3d/static/xlsx/'+username+logintime+'.xlsx'
            fs = FileSystemStorage(location='plot3d/static/xlsx/')
            if (fs.exists(username+logintime+'.xlsx')):
                print('file_exist')
                wb = openpyxl.load_workbook(dest_filename)
                sheets = wb.get_sheet_names()
                ddata = []
                selectx = ''
                selecty = ''
                columns = []
                sheet = wb.active
                for j in range(1, sheet.max_column + 1):
                    columns.append(get_column_letter(j))
                    selecty += '<option value="' + str(j) + '">Column' + str(j) + '</option>'
                    selectx += '<option value="' + str(j) + '">Column' + str(j) + '</option>'

                print(columns)
                print(selectx)
                print(selecty)
                for s in sheets:
                    sheet = wb.get_sheet_by_name(s)

                    x = []
                    print(sheet.max_row,sheet.max_column)
                    for i in range(1, sheet.max_row + 1):
                        dtemp = []
                        for j in range(1, sheet.max_column + 1):
                            dtemp.append((sheet.cell(row=i, column=j).value))
                        ddata.append(dtemp)

                    # print(ddata)
                    # print('hellooo')
                    ddata = str(ddata)
                    ddata = ddata[1:-1]
                tabled = ddata
                return render(request, '3d_index.html', locals())
            elif 'user_name' in request.COOKIES:

                username = request.COOKIES['user_name']

                ddata = []
                for i in range(1, 10):
                    dtemp = []
                    for j in range(1, 3):
                        dtemp.append("")
                    ddata.append(dtemp)

                    # print(ddata)
                    # print('hellooo')
                ddata = str(ddata)
                ddata = ddata[1:-1]
                tabled = ddata

                return render(request,'3d_index.html',locals())
            else:
                return redirect('./../?usr=001', 'default.html', request)
        else:
            return redirect('./../?usr=001', 'default.html', request)
    elif(request.POST.get('fsubmit')):
        print('post fsubmit  data')

        username = 'InvalidUser'
        logintime = '0'
        if 'user_name' in request.COOKIES:
            username = request.COOKIES['user_name']
        if 'login_time' in request.COOKIES:
            logintime = request.COOKIES.get('login_time')
        fs = FileSystemStorage(location='plot3d/static/xlsx/')
        fn = username + logintime + '.xlsx'
        if (fs.exists(username + logintime + '.xlsx')):
            os.remove('plot3d/static/xlsx/' + username + logintime + '.xlsx')

        fs.save(fn, request.FILES.get("file"))
        wb = openpyxl.load_workbook(request.FILES.get("file"))
        sheets = wb.get_sheet_names()
        ddata=[]
        selectx = ''
        selecty = ''
        columns=[]
        sheet=wb.active
        for j in range(1,sheet.max_column+1):
            columns.append(get_column_letter(j))
            selecty+='<option value="'+str(j)+'">Column'+str(j)+'</option>'
            selectx += '<option value="' + str(j) + '">Column' + str(j) + '</option>'

        print(columns)
        print(selectx)
        print(selecty)
        for s in sheets:
            sheet = wb.get_sheet_by_name(s)

            x = []

            for i in range(1, sheet.max_row + 1):
                dtemp = []
                for j in range(1, sheet.max_column + 1):
                    dtemp.append((sheet.cell(row=i, column=j).value))
                ddata.append(dtemp)

            # print(ddata)
            # print('hellooo')
            ddata = str(ddata)
            ddata = ddata[1:-1]
        tabled=ddata


        return render(request,'3d_index.html',locals())
    elif(request.POST.get('funsubmit')):
       return redirect('./../fplot','findex.html',request)
    elif(request.POST.get('funpsubmit')):
        return redirect('./../pplot', 'pindex.html', request)
    elif (request.POST.get('annsubmit')):
        return redirect('./../ann', 'ann_index.html', request)
    elif(request.POST.get('bsubmit')):

        username = 'InvalidUser'
        logintime = '0'
        if 'user_name' in request.COOKIES:
            username = request.COOKIES['user_name']
        if 'login_time' in request.COOKIES:
            logintime = request.COOKIES.get('login_time')


        ddata=[]
        for i in range(1, 10):
            dtemp = []
            for j in range(1, 3):
                dtemp.append("")
            ddata.append(dtemp)

            #print(ddata)
            #print('hellooo')
        ddata = str(ddata)
        ddata = ddata[1:-1]
        tabled=ddata
        return render(request,'3d_index.html',locals())
    elif(request.POST.get('plot') and 'user_name' in request.COOKIES) :
        print('post plot')

        username = 'not a valid user'
        if 'user_name' in request.COOKIES:
            username = request.COOKIES['user_name']
        logintime = '0'
        if 'login_time' in request.COOKIES:
            logintime = request.COOKIES.get('login_time')

        plot3dtype = request.POST.get('plot3dtype')

        name=request.POST.get("name")
        xlable=request.POST.get("xlable")
        ylable=request.POST.get("ylable")
        zlable = request.POST.get("zlable")

        countcols=request.POST.get('countcols')

        print(countcols)
        fig = plt.figure(figsize=(7,5))
        ax = fig.add_subplot(111, projection='3d')
        fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold',color='red')

        fig.subplots_adjust(top=0.85)
        ax = fig.gca(projection='3d')

        plt.grid(True)

        maxx=0
        maxy=0
        minx=0
        miny=0
        ddata=[]

        TOOLS = 'box_select,resize,reset,pan,tap,wheel_zoom,save,hover'

        bplot=figure(title=request.POST.get('name'),plot_width=640,plot_height=480,x_axis_label=request.POST.get('xlable'),y_axis_label=request.POST.get('ylable'),tools=TOOLS, toolbar_location='above')
        bplot.title.text_color="red"
        bplot.title.text_font_size='22px'



        ycolumns=[]
        xcolumns=[]
        zcolums=[]
        x=[]
        y=[]
        z=[]

        dest_filename = 'plot3d/static/xlsx/' + username + logintime + '.xlsx'
        wb = openpyxl.load_workbook(dest_filename)
        sheets = wb.get_sheet_names()

        for s in sheets:
            sheet=wb.get_sheet_by_name(s)
            for i in range(1,sheet.max_row+1):
                dtemp=[]
                for j in range(1,sheet.max_column+1):
                    if plot3dtype!='bar' and plot3dtype!='pie':
                        dtemp.append(float(sheet.cell(row=i,column=j).value))
                    else:
                        dtemp.append((sheet.cell(row=i, column=j).value))
                ddata.append(dtemp)

        #print(ddata)
        #print('hellooo')
        ddata=str(ddata)
        ddata=ddata[1:-1]


        plot3dtype = request.POST.get('plot3dtype')
        print(plot3dtype)
        ''' plotting data are saved in x and y[]'''
        div_script=''
        div_tag=''
        xcol=request.POST.get('xdata')
        ycol = request.POST.get('ydata')
        zcol = request.POST.get('zdata')
        zcolnum=request.POST.get('zdatanum')
        colr=request.POST.get('color')
        if plot3dtype!='bar' and plot3dtype!='pie':
            for i in range(1,sheet.max_row+1):
                x.append(float(sheet.cell(row=i,column=int(xcol)).value))



        '''get selected columns data'''
        if plot3dtype!='histogram':
            for i in range(1, sheet.max_row+1):
                y.append(float(sheet.cell(row=i,column=int(ycol)).value))


        print(plot3dtype)
        if plot3dtype=='line':
            print('line')
            x=[]
            y=[]
            z=[]
            for i in range(1,sheet.max_row+1):
                x.append(float(sheet.cell(row=i,column=int(xcol)).value))


            for i in range(1, sheet.max_row+1):
                y.append(float(sheet.cell(row=i,column=int(ycol)).value))

            if zcolnum:
                z=int(zcolnum)
            else:
                for i in range(1, sheet.max_row+1):
                    z.append(float(sheet.cell(row=i,column=int(zcol)).value))
            print(z)

            ax.plot(xs=x, ys=y,zs=z, marker='.',color=colr)


        elif plot3dtype=='bar':
            print('bar')

            x = []
            y = []
            z = []
            for i in range(1, sheet.max_row + 1):
                x.append((sheet.cell(row=i, column=int(xcol)).value))

            for i in range(1, sheet.max_row + 1):
                y.append(float(sheet.cell(row=i, column=int(ycol)).value))

            if zcolnum:
                z = int(zcolnum)
            else:
                for i in range(1, sheet.max_row + 1):
                    z.append(float(sheet.cell(row=i, column=int(zcol)).value))


            for zs in z:
                ax.bar(x, y,zs, color=colr,zdir='y',alpha=0.7)

            dic = {'values': y, 'names': x}
            df = pd.DataFrame(dic)
            bar = Bar(df, 'names', values='values', color=colr, legend='top_left')
            div_script, div_tag = autoload_static(bar, CDN, "js")
        elif plot3dtype=='scatter':
           print('scatter')
           x = []
           y = []
           z = []
           for i in range(1, sheet.max_row + 1):
               x.append(float(sheet.cell(row=i, column=int(xcol)).value))

           for i in range(1, sheet.max_row + 1):
               y.append(float(sheet.cell(row=i, column=int(ycol)).value))

           if zcolnum:
               z = int(zcolnum)
           else:
               for i in range(1, sheet.max_row + 1):
                   z.append(float(sheet.cell(row=i, column=int(zcol)).value))

           ax.scatter(x,y,z)

           dic = {ylable: y, xlable: x}
           df = pd.DataFrame(dic)
           scatr = Scatter(df, x=xlable, y=ylable, color=colr)
           div_script, div_tag = autoload_static(scatr, CDN, "js")

        elif plot3dtype=='contour':
            z=[]
            zcol=request.POST.get('zdata')
            for i in range(1, sheet.max_row + 1):
                z.append(float(sheet.cell(row=i, column=int(zcol)).value))

            Z=[]
            for i in range(0,len(y[0])):
                Z.append(z)
            for j in range(0,len(y)):
                plt.contour(x, y[j],Z)

        elif plot3dtype=='histogram':
            print('histogram')
            ax.hist3d(x,y,color='#2bb3ff')
            print(x)

            dic = { xlable: x}
            df = pd.DataFrame(dic)
            print(df)
            hist= Histogram(df, values=xlable, color='blue')
            div_script, div_tag = autoload_static(hist, CDN, "js")
        else:
            print('else')
            for j in range(0,len(y)) :
                plt.plot(x,y[j],marker='.', markersize=2, alpha=1, color=colorscol[j], label="Fun" + str(j))
                bplot.line(x,y[j],color=colorscol[j],legend='Fun'+str(j))
                if(i==sheet.max_row-1):
                    plt.annotate('fun-'+str(j-1), xy=(float(sheet.cell(row=i, column=1).value), float(sheet.cell(row=i, column=j).value)),xytext=(float(sheet.cell(row=i, column=1).value-(maxx/6)), float(sheet.cell(row=i, column=j).value)),
                            arrowprops=dict(facecolor='black',arrowstyle='->'),horizontalalignment='left',verticalalignment='bottom'
                                         )

        plt.legend(loc=0)
        fig1 = plt.gcf()
        if(FileSystemStorage('/static/img/').exists(username+logintime+'.png')):
            os.remove('static/img/'+username+logintime+'.png')
        plt.savefig('plot3d/static/img/'+str(username)+str(logintime)+'.png', dpi=200,bbox_inches = 'tight')
        plt.savefig('plot3d/static/img/' + username + logintime + '.jpg', dpi=200)
        fig1.savefig('plot3d/static/img/' + username + logintime + '.eps', dpi=200)
        fig1.savefig('plot3d/static/img/' + username + logintime + '.pdf', dpi=200)
        fig1.savefig('plot3d/static/img/' + username + logintime + '.svg', dpi=200)
        fig1.savefig('plot3d/static/img/' + username + logintime + '.tiff', dpi=200)
        fig1.savefig('plot3d/static/img/' + username + logintime + '.ps', dpi=200)
        fig1.savefig('plot3d/static/img/' + username + logintime + '.svgz', dpi=200)
        # fig1.savefig('plot3d/static/img/test.pgf', dpi=200)#latex not installed sudo apt-get install texlive-xetex
        fig1.savefig('plot3d/static/img/' + username + logintime + '.jpeg', dpi=200)
        # fig1.savefig('plot3d/static/img/test.raw', dpi=200)#if needed
        fig1.savefig('plot3d/static/img/' + username + logintime + '.tif', dpi=200)

        saved = True

        dis = 'plot3d/static/img/' + username+logintime + '.png'
        #shutil.copyfile('plot3d/static/img/'+username+logintime+'.png', dis)
        saved = True
        imagelink = 'img/' +username+logintime  + '.png'
        filelink = 'xlsx/'+username+logintime+'.xlsx'

        bplot.legend.location = "top_left"
        bplot.legend.background_fill_color = "white"
        bplot.legend.background_fill_alpha = 0.5


        image = 'img/' + username + logintime
        plotplot3d=1
        tabled=1
        #print(div_tag,div_script)
        username='InvalidUser'
        logintime='0'
        if 'user_name' in request.COOKIES:
                username = request.COOKIES['user_name']
        if 'login_time' in request.COOKIES:
                logintime=request.COOKIES['login_time']

        dest_filename = 'plot3d/static/xlsx/' + username + logintime + '.xlsx'
        wb = openpyxl.load_workbook(dest_filename)
        sheets = wb.get_sheet_names()

        selectx = ''
        selecty = ''
        columns = []
        for s in sheets:
            sheet=wb.get_sheet_by_name(s)
            for j in range(1, sheet.max_column + 1):
                columns.append(get_column_letter(j))
                selecty += '<option value="' + str(j) + '">Column' + str(j) + '</option>'
                selectx += '<option value="' + str(j) + '">Column' + str(j) + '</option>'

        print(columns)
        print(selectx)
        print(selecty)

        print(div_tag)
        print(div_script)
        return render(request,'3d_index.html', locals())
    else:
        print('nothing')
        ddata = 0
        tabled = 0
        plotplot3d = 0
        username='InvalidUser'
        return redirect('./../?usr=0','3d_index.html',request)


def interpolation(request):
    if request.method != 'POST':
        return render(request, 'interpolation.html', locals())
    else:
        msg = ""
        msg = "post"
        msg += request.POST.get("name")
        fs = FileSystemStorage(location='plot3d/static/xlsx/')
        fs._save("interpolation.xlsx", request.FILES.get("file"))

        fig = plt.figure()
        fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold', color='red')
        ax = fig.add_subplot(111)

        fig.subplots_adjust(top=0.85)
        ax.set_title(request.POST.get('axis'))
        ax.set_xlabel(request.POST.get('xlable'))
        ax.set_ylabel(request.POST.get('ylable'))
        ax.plot()
        plt.grid(True)

        wb = openpyxl.load_workbook(request.FILES.get("file"))
        sheets = wb.get_sheet_names()

        col = []
        colr = ['r', 'b', 'g', 'y']

        for s in sheets:
            sheet = wb.get_sheet_by_name(s)
            msg += str(sheet.max_column)
            maxx = 0
            minx = 999999999
            x = []
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=1).value))
                if (x[i - 1] > maxx):
                    maxx = x[i - 1]
                if (x[i - 1] < minx):
                    minx = x[i - 1]

            y = []
            for j in range(2, sheet.max_column + 1):
                y_tmp = []
                for i in range(1, sheet.max_row + 1):
                    y_tmp.append(float(sheet.cell(row=i, column=j).value))
                y.append(y_tmp)

            for j in range(0, len(y)):
                f = interp1d(x, y[j])
                f2 = interp1d(x, y[j], kind='cubic')
                xnew = np.linspace(minx, maxx, num=1000, endpoint=True)
                plt.plot(x, y[j], 'o', xnew, f(xnew), '-', xnew, f2(xnew), '--')

                msg += 'sinx'
        plt.legend(loc=0)
        fig1 = plt.gcf()

        if (FileSystemStorage('/static/img/').exists('interpolation.png')):
            os.remove('static/img/interpolation.png')

        fig1.savefig('plot3d/static/img/interpolation.png', dpi=100)
        saved = True

        msg += "post and saved"

        return render(request, 'interpolation.html', locals())














def curveFit(request):
    def fitfun(x,b):
        return  np.exp(-b*x)
    def fitfun1(x,a,b):
        return x*a+b
    def fitfun2(x,a,b,c):
        return a*x*x+b*x+c
    def fitfun3(x,a,b,c,d):
        return a*x*x*x+b*x*x+c*x+d
    def fitfun4(x,a,b,c):
        return a*np.cos(2*np.pi*x+b)+c


    if request.method!='POST':
        return render(request, 'curvefitting.html', locals())
    else:
        msg = ""
        msg = "post"
        msg += request.POST.get("name")
        if mydocfrom.is_valid():

            '''
            #fake data createing
            x=np.linspace(0,4,50)
            temp=fitfun(x,2.5,1.3,0.5)
            y=temp+0.25*np.random.normal(size=len(temp))
            #scipy optimise module contains a least square curve fit rutine
            '''


            fs = FileSystemStorage(location='plot3d/static/xlsx/')
            fs._save("curve_fit.xlsx", request.FILES.get("file"))

            fig = plt.figure()
            fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold', color='red')
            ax = fig.add_subplot(111)

            fig.subplots_adjust(top=0.85)
            ax.set_title(request.POST.get('axis'))
            ax.set_xlabel(request.POST.get('xlable'))
            ax.set_ylabel(request.POST.get('ylable'))
            ax.plot()
            plt.grid(True)

            wb = openpyxl.load_workbook(request.FILES.get("file"))
            sheets = wb.get_sheet_names()

            col = []
            colr = ['r', 'b', 'g', 'y','c','m','k']
            maxx = 0
            maxy = 0
            minx = 0
            miny = 0

            y=[]
            x=[]
            for s in sheets:
                sheet = wb.get_sheet_by_name(s)
                msg += str(sheet.max_column)



                #getting x values
                for i in range(1,sheet.max_row+1):
                    x.append(float(sheet.cell(row=i,column=1).value))

                #getting y values:
                x=np.array(x)

                for j in range(2, sheet.max_column + 1):
                    y_temp = []
                    for i in range(1, sheet.max_row+1):
                        y_temp.append( float(sheet.cell(row=i, column=j).value))
                    y.append(y_temp)


                print (y)
                for y_val in y:
                    y_val=np.array(y_val)
                    if (int(request.POST.get('variable')) == 2):
                        fitParams, fitCovariances = curve_fit(fitfun1, x, y_val)
                        #msg += str(fitfun1(x, fitParams[0], fitParams[1]))
                        fit=fitfun1(x, fitParams[0], fitParams[1])
                        line1 = plt.plot(x, fitfun1(x, fitParams[0], fitParams[1]), label="Line 1")
                        err1 = plt.errorbar(x, y_val, fmt='bs', yerr=0.1, label="Error-1")
                    elif(int(request.POST.get('variable'))==3):
                        fitParams, fitCovariances = curve_fit(fitfun2, x, y_val)
                        #msg += str(fitfun2(x, fitParams[0], fitParams[1], fitParams[2]))
                        line2 = plt.plot(x, fitfun2(x, fitParams[0], fitParams[1], fitParams[2]), label="Line 2")
                        err2 = plt.errorbar(x, y_val, fmt='g-.', yerr=0.1, label="Error-2")
                    elif(int(request.POST.get('variable'))==4):
                        fitParams, fitCovariances = curve_fit(fitfun3, x, y_val)
                        #msg += str(fitfun3(x, fitParams[0], fitParams[1], fitParams[2],fitParams[3]))
                        line3 = plt.plot(x, fitfun3(x, fitParams[0], fitParams[1], fitParams[2],fitParams[3]), label="Line 3")
                        err3 = plt.errorbar(x, y_val, fmt='b^', yerr=0.1, label="Error-3")
                    elif((int(request.POST.get('variable'))==5)):
                        fitParams, fitCovariances = curve_fit(fitfun4, x, y_val)
                        # msg += str(fitfun3(x, fitParams[0], fitParams[1], fitParams[2],fitParams[3]))
                        err4 = plt.errorbar(x, y_val, fmt='bo', yerr=2.0, label="Error-4")

                        line4 = plt.plot(x, fitfun4(x, fitParams[0], fitParams[1], fitParams[2]),
                                         label="Line 4",color='r',marker='^',markersize=3.0)


                        #plt.ylabel('Temperature (C)', fontsize=16)
                    #plt.xlabel('Time', fontsize=16)
                    #plt.xlim(0, 4.1)

                    # plot data as red square with vertical errors
                    # plot best fit plot3d

                fig1 = plt.gcf()
                plt.legend()#to plot top side square for plot detail
                if (FileSystemStorage('/static/img/').exists('curve_fit.png')):
                    os.remove('static/img/curve_fit.png')

                fig1.savefig('plot3d/static/img/curve_fit.png', dpi=200)
                saved = True
                return render(request, 'curvefitting.html', locals())



def lineFit(request):
    def fitfun(x,b):
        return  np.exp(-b*x)
    def fitfun1(x,a,b):
        return x*a+b
    def fitfun2(x,a,b,c):
        return a*x*x+b*x+c
    def fitfun3(x,a,b,c,d):
        return a*x*x*x+b*x*x+c*x+d
    def fitfun4(x,a,b,c):
        return a*np.cos(2*np.pi*x+b)+c


    if request.method!='POST':
        return render(request, 'linefitting.html', locals())
    else:
        msg = ""
        msg = "post"
        msg += request.POST.get("name")

        '''
        #fake data createing
        x=np.linspace(0,4,50)
        temp=fitfun(x,2.5,1.3,0.5)
        y=temp+0.25*np.random.normal(size=len(temp))
        #scipy optimise module contains a least square curve fit rutine
        '''


        fs = FileSystemStorage(location='plot3d/static/xlsx/')
        fs._save("line_fit.xlsx", request.FILES.get("file"))

        fig = plt.figure()
        fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold', color='red')
        ax = fig.add_subplot(111)

        fig.subplots_adjust(top=0.85)
        ax.set_title(request.POST.get('axis'))
        ax.set_xlabel(request.POST.get('xlable'))
        ax.set_ylabel(request.POST.get('ylable'))

        plt.grid(True)

        wb = openpyxl.load_workbook(request.FILES.get("file"))
        sheets = wb.get_sheet_names()

        col = []
        colr = ['r', 'b', 'g', 'y', 'c', 'm', 'k']
        maxx = 0
        maxy = 0
        minx = 0
        miny = 0

        y = []
        x = []
        for s in sheets:
            sheet = wb.get_sheet_by_name(s)
            msg += str(sheet.max_column)

            # getting x values
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=1).value))

            # getting y values:
            x = np.array(x)

            for j in range(2, sheet.max_column + 1):
                y_temp = []
                for i in range(1, sheet.max_row + 1):
                    y_temp.append(float(sheet.cell(row=i, column=j).value))
                y.append(y_temp)

            print(y)
            for y_val in y:
                y_val = np.array(y_val)
                if (int(request.POST.get('variable')) == 2):
                    # parameters
                    a = 0.8
                    b = -4
                    x_val = polyval([a, b], y_val)

                    # Linear regressison -polyfit - polyfit can be used other orders polys
                    (ar, br) = polyfit(x, x_val, 1)
                    x_reg = np.polyval([ar, br], x)
                    # compute the mean square error
                    err = sqrt(sum((x_reg - x_val) ** 2) / i)

                    # matplotlib ploting
                    plt.title('Linear Regression Example')
                    plt.plot(x, x_val, 'g.--')
                    plt.plot(x, x_reg, 'r.-')

                    # plt.show()

                    # Linear regression using stats.linregress
                    slope, intercept, r_value, p_value, std_err = stats.linregress(x, x_val)
                    y_pre = intercept + slope * x
                    plt.plot(x, y_pre, 'b.-')

                    plt.legend(['original', 'regression', 'regression_using_fn'])

                elif (int(request.POST.get('variable')) == 3):
                    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y_val)
                    y_pre = intercept + slope * x
                    plt.plot(x, y_pre, 'b.-')

                    plt.legend(['original', 'regression', 'regression_using_fn'])

                    err1 = plt.errorbar(x, y_val, fmt='bs', yerr=0.1, label="Error-3")
                else:
                    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y_val)
                    y_pre = intercept + slope * x
                    plt.plot(x, y_pre, 'b.-')

                    plt.legend(['original', 'regression', 'regression_using_fn'])

                    err1 = plt.errorbar(x, y_val, fmt='bs', yerr=0.1, label="Error-n")
                    # plt.xlabel('Time', fontsize=16
                    # plt.xlim(0, 4.1)

                    # plot data as red square with vertical errors
                    # plot best fit plot3d

            fig1 = plt.gcf()
            # plt.legend()#to plot top side square for plot detail
            if (FileSystemStorage('/static/img/').exists('line_fit.png')):
                os.remove('static/img/line_fit.png')

            fig1.savefig('plot3d/static/img/line_fit.png', dpi=1000)
            saved = True
            return render(request, 'linefitting.html', locals())
