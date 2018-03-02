from django.shortcuts import render,render_to_response,redirect

from django.template import RequestContext
from django.core.files.storage import FileSystemStorage

import  os
import openpyxl
from openpyxl import  Workbook
from openpyxl.utils import get_column_letter
from origin import settings
from django.core.files import File
from matplotlib import pyplot as plt
import numpy as np
from scipy import linspace, polyval, polyfit, sqrt, stats
from scipy.optimize import curve_fit
import matplotlib.mlab as mlab
from scipy.interpolate import  interp1d
# Create your views here.
from django.http import HttpResponse,HttpResponseRedirect,HttpRequest
import  pandas as pd
from django.views.decorators.csrf import csrf_exempt
import shutil
from bokeh.plotting import figure, output_file, show
from bokeh.resources import CDN
from bokeh.embed import components
from bokeh.embed import autoload_static
from bokeh.charts import Bar,BoxPlot,Area,Histogram,Line,Scatter,TimeSeries,Donut
import datetime,time
import glob
from functools import  reduce
from django.utils.html import escape
from itertools import cycle
from sklearn.cluster import AffinityPropagation,KMeans
from sklearn import metrics
from sklearn.datasets.samples_generator import make_blobs
from matplotlib.mlab import griddata
from sklearn.linear_model import LinearRegression
from scipy.stats.stats import pearsonr
from scipy.interpolate import interp1d,interpnd,interp2d

@csrf_exempt
def save_data(request):
    print("save_data")
    if request.method=='POST':
        rows=request.POST.get('rows')
        cols=request.POST.get('cols')
        #print(rows,cols)
        rows=int(rows)
        cols=int(cols)
        index_table(request,request.POST.get('data'),rows,cols)
        return HttpResponse(1)
    return HttpResponse(0)

def myfloat(seq,row,col):
    print('myfloat')
    ans=[]

    temp = []
    i=0
    j=0
    print(row,col)
    for x in seq:
        try:
            if col>j:
                if len(x)>0:
                    temp.append((x))
                    j+=1
                    print(x,j)
            else:
                j=1
                print('else')
                print(ans)
                print(x,j)
                ans.append(temp)
                print(ans)
                temp=[]
                if len(x)>0:
                    temp.append((x))
        except ValueError:
            print('value error')
            print(len(temp))
            print(temp)
            if len(temp)>0:
                ans.append(temp)
                print(temp)
                print(ans)
            j=0
            temp=[]
    if len(temp)>0:
        ans.append(temp)
    print(ans)
    return ans

def index_table(request,ddata,row,col):
    print('index table')
    rows=int(row)
    cols=int(col)

    username=request.COOKIES.get('user_name')
    logintime=request.COOKIES.get('login_time')
    print(username)
    print(logintime)
    dest_filename = 'analysis/static/xlsx/'+username+logintime+'.xlsx'
    fs = FileSystemStorage(location='analysis/static/xlsx/')
    if (fs.exists(username+logintime+'.xlsx')):
        os.remove('analysis/static/xlsx/'+username+logintime+'.xlsx')

    print(ddata)
    data=ddata.split(",")
    print(data)
    data=myfloat(data,rows,cols)
    print(data)
    #data=list(data)
    #print(data)
    print('data')
    k=0
    l=0
    print(cols,rows)
    wb = Workbook()
    ws = wb.active
    print(len(data))
    print(len(data[0]))
    try:
        for i in range(0,len(data)):

                for j in range(0,len(data[i])+1):
                    try:
                        ws.cell(column=j+1, row=i+1,value=data[i][j])
                    except:
                        print('break')
                        break

        wb.save(dest_filename)
        print("saved")
    except:
        wb.save(dest_filename)
        print('saved with exception')


@csrf_exempt
def del_user(request):
    request.session.set_test_cookie()
    ddata = 0
    tabled = 0
    plotanalysis = 0
    username='not a valid user'
    logintime=0
    if 'user_name' in request.COOKIES:
        username=request.COOKIES.get('user_name')
    if 'login_time' in request.COOKIES:
        logintime=request.COOKIES.get('login_time')

    pathi='analysis/static/img/'+username+'*.*'
    pathf = 'analysis/static/xlsx/' + username + str(logintime) + '*.*'

    for hgx in glob.glob(pathi):
        os.remove(hgx)
    for hgx in glob.glob(pathf):
        os.remove(hgx)

    if request.session.test_cookie_worked():
        request.session.delete_test_cookie()
        response = HttpResponse("Bye")
        response.delete_cookie('user_name')
        response.delete_cookie('login_time')
        return  response
    else:
        return HttpResponse("Sorry")

@csrf_exempt
def default(request):
    if 'user_name' in request.COOKIES:
        username=request.COOKIES['user_name']
        return redirect('./analysis','module.html',locals())
    elif request.method=='POST':
        username=request.POST.get('username')
        response=redirect('./analysis','index.html',request)
        response.set_cookie('user_name',request.POST.get('username'))
        response.set_cookie('login_time',time.time())
        #return redirect('./analysis',request,locals())
        return response
    else:
        return render(request,'default.html',locals())












@csrf_exempt
def index(request):
    errormsg=''
    if request.method!='POST':
        print('not post')
        ddata=0
        tabled=0
        plotanalysis=0
        if 'user_name' in request.COOKIES and 'login_time' in request.COOKIES:
            username = request.COOKIES['user_name']
            logintime=request.COOKIES.get('login_time')
            dest_filename = 'analysis/static/xlsx/'+username+logintime+'.xlsx'
            fs = FileSystemStorage(location='analysis/static/xlsx/')
            if (fs.exists(username+logintime+'.xlsx')):
                print('file_exist')
                wb = openpyxl.load_workbook(dest_filename)
                sheets = wb.get_sheet_names()
                ddata = []
                selectx = ''
                selecty = ''
                columns = []
                sheet = wb.active
                for j in range(1, sheet.max_column + 1):
                    columns.append(get_column_letter(j))
                    selecty += '<option value="' + str(j) + '">Column' + str(j) + '</option>'
                    selectx += '<option value="' + str(j) + '">Column' + str(j) + '</option>'

                print(columns)
                print(selectx)
                print(selecty)
                for s in sheets:
                    sheet = wb.get_sheet_by_name(s)

                    x = []
                    print(sheet.max_row,sheet.max_column)
                    for i in range(1, sheet.max_row + 1):
                        dtemp = []
                        for j in range(1, sheet.max_column + 1):
                            dtemp.append((sheet.cell(row=i, column=j).value))
                        ddata.append(dtemp)

                    # print(ddata)
                    # print('hellooo')
                    ddata = str(ddata)
                    ddata = ddata[1:-1]
                tabled = ddata
                return render(request, 'analysis_index.html', locals())
            elif 'user_name' in request.COOKIES:


                username = request.COOKIES['user_name']

                ddata = []
                for i in range(1, 10):
                    dtemp = []
                    for j in range(1, 3):
                        dtemp.append("")
                    ddata.append(dtemp)

                    # print(ddata)
                    # print('hellooo')
                ddata = str(ddata)
                ddata = ddata[1:-1]
                tabled = ddata

                return render(request,'analysis_index.html',locals())
            else:
                return redirect('./../?usr=001', 'default.html', request)
        else:
            return redirect('./../?usr=001', 'default.html', request)
    elif(request.POST.get('fsubmit')):
        print('post fsubmit  data')

        username = 'InvalidUser'
        logintime = '0'
        if 'user_name' in request.COOKIES:
            username = request.COOKIES['user_name']
        if 'login_time' in request.COOKIES:
            logintime = request.COOKIES.get('login_time')
        fs = FileSystemStorage(location='analysis/static/xlsx/')
        fn = username + logintime + '.xlsx'
        if (fs.exists(username + logintime + '.xlsx')):
            os.remove('analysis/static/xlsx/' + username + logintime + '.xlsx')

        fs.save(fn, request.FILES.get("file"))
        wb = openpyxl.load_workbook(request.FILES.get("file"))
        sheets = wb.get_sheet_names()
        ddata=[]
        selectx = ''
        selecty = ''
        columns=[]
        sheet=wb.active
        for j in range(1,sheet.max_column+1):
            columns.append(get_column_letter(j))
            selecty+='<option value="'+str(j)+'">Column'+str(j)+'</option>'
            selectx += '<option value="' + str(j) + '">Column' + str(j) + '</option>'

        print(columns)
        print(selectx)
        print(selecty)
        for s in sheets:
            sheet = wb.get_sheet_by_name(s)

            x = []

            for i in range(1, sheet.max_row + 1):
                dtemp = []
                for j in range(1, sheet.max_column + 1):
                    dtemp.append((sheet.cell(row=i, column=j).value))
                ddata.append(dtemp)

            # print(ddata)
            # print('hellooo')
            ddata = str(ddata)
            ddata = ddata[1:-1]
        tabled=ddata


        return render(request,'analysis_index.html',locals())
    elif(request.POST.get('funsubmit')):
       return redirect('./../fplot','findex.html',request)
    elif(request.POST.get('funpsubmit')):
        return redirect('./../pplot', 'pindex.html', request)
    elif (request.POST.get('annsubmit')):
        return redirect('./../ann', 'ann_index.html', request)
    elif(request.POST.get('bsubmit')):

        username = 'InvalidUser'
        logintime = '0'
        if 'user_name' in request.COOKIES:
            username = request.COOKIES['user_name']
        if 'login_time' in request.COOKIES:
            logintime = request.COOKIES.get('login_time')


        ddata=[]
        for i in range(1, 10):
            dtemp = []
            for j in range(1, 3):
                dtemp.append("")
            ddata.append(dtemp)

            #print(ddata)
            #print('hellooo')
        ddata = str(ddata)
        ddata = ddata[1:-1]
        tabled=ddata
        return render(request,'analysis_index.html',locals())
    elif(request.POST.get('plotdata')):
        print('post plot data')
        name = request.POST.get("name")
        xlable = request.POST.get("xlable")
        ylable = request.POST.get("ylable")

        fig = plt.figure()
        fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold', color='red')
        ax = fig.add_subplot(111)

        fig.subplots_adjust(top=0.85)
        ax.set_xlabel(request.POST.get('xlable'))
        ax.set_ylabel(request.POST.get('ylable'))
        ax.plot()
        plt.grid(True)
        username='InvalidUser'
        logintime='0'
        if 'user_name' in request.COOKIES:
            username = request.COOKIES['user_name']
        if 'login_time' in request.COOKIES:
            logintime = request.COOKIES.get('login_time')


        col = []

        colr = ['r', 'b', 'g', 'y']
        maxx = 0
        maxy = 0
        minx = 0
        miny = 0
        ddata = []
        msg = ''
        wb = openpyxl.load_workbook('analysis/static/xlsx/' + username + logintime + '.xlsx')
        sheets = wb.get_sheet_names()

        for s in sheets:
            sheet = wb.get_sheet_by_name(s)
            msg += str(sheet.max_column)
            x = []

            for i in range(1, sheet.max_row + 1):
                dtemp = []
                for j in range(1, sheet.max_column + 1):
                    dtemp.append(float(sheet.cell(row=i, column=j).value))
                ddata.append(dtemp)

            #print(ddata)
            #print('hellooo')
            ddata = str(ddata)
            ddata = ddata[1:-1]

            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=1).value))

            y = []
            for j in range(2, sheet.max_column + 1):
                y_tmp = []
                for i in range(1, sheet.max_row + 1):
                    y_tmp.append(float(sheet.cell(row=i, column=j).value))
                y.append(y_tmp)

            bcolr = ['red', 'blue', 'green', 'yellow']
            TOOLS = 'box_select,resize,reset,pan,tap,wheel_zoom,save,hover'


            bplot=figure(title=request.POST.get('name'),plot_width=640,plot_height=480,x_axis_label=request.POST.get('xlable'),y_axis_label=request.POST.get('ylable'),tools=TOOLS, toolbar_location='above')
            bplot.title.text_color="red"
            bplot.title.text_font_size='22px'




            for j in range(0, len(y)):
                plt.plot(x, y[j], marker='.', markersize=2, alpha=1, color=colr[(j % 4)], label="Fun" + str(j))
                bplot.line(x, y[j],color=bcolr[(j % 4)], legend='Fun' + str(j))

                if (i == sheet.max_row - 1):
                    plt.annotate('fun-' + str(j - 1),
                                 xy=(
                                 float(sheet.cell(row=i, column=1).value), float(sheet.cell(row=i, column=j).value)),
                                 xytext=(float(sheet.cell(row=i, column=1).value - (maxx / 6)),
                                         float(sheet.cell(row=i, column=j).value)),
                                 arrowprops=dict(facecolor='black', arrowstyle='->'), horizontalalignment='left',
                                 verticalalignment='bottom'
                                 )

        plt.legend(loc=0)
        fig1 = plt.gcf()

        if (FileSystemStorage('/static/img/').exists('test.png')):
            os.remove('static/img/test.png')

        fig1.savefig('analysis/static/img/'+username+logintime+'.png', dpi=200)
        saved = True



        fig1.savefig('analysis/static/img/'+username+logintime+'.jpg', dpi=200)
        fig1.savefig('analysis/static/img/'+username+logintime+'.eps', dpi=200)
        fig1.savefig('analysis/static/img/'+username+logintime+'.pdf', dpi=200)
        fig1.savefig('analysis/static/img/'+username+logintime+'.svg', dpi=200)
        fig1.savefig('analysis/static/img/'+username+logintime+'.tiff', dpi=200)
        fig1.savefig('analysis/static/img/'+username+logintime+'.ps', dpi=200)
        fig1.savefig('analysis/static/img/'+username+logintime+'.svgz', dpi=200)
        #fig1.savefig('analysis/static/img/test.pgf', dpi=200)#latex not installed sudo apt-get install texlive-xetex
        fig1.savefig('analysis/static/img/'+username+logintime+'.jpeg', dpi=200)
        #fig1.savefig('analysis/static/img/test.raw', dpi=200)#if needed
        fig1.savefig('analysis/static/img/'+username+logintime+'.tif', dpi=200)

        dis = 'analysis/static/img/'+username+logintime+'.png'
        saved = True
        imagelink = 'img/'+username+logintime+'.png'
        filelink = 'xlsx/'+username+logintime+'.xlsx'
        bplot.legend.location = "top_left"
        bplot.legend.background_fill_color = "white"
        bplot.legend.background_fill_alpha = 0.5
        div_script, div_tag = autoload_static(bplot, CDN, "js")
        image='img/'+username+logintime
        plotanalysis=1
        tabled=1
        print(div_script)
        print(div_tag)
        username = 'InvalidUser'
        logintime = '0'
        if 'user_name' in request.COOKIES:
            username = request.COOKIES['user_name']
        if 'login_time' in request.COOKIES:
            logintime = request.COOKIES['login_time']

        dest_filename = 'analysis/static/xlsx/' + username + logintime + '.xlsx'
        wb = openpyxl.load_workbook(dest_filename)
        sheets = wb.get_sheet_names()

        selectx = ''
        selecty = ''
        columns = []
        for s in sheets:
            sheet = wb.get_sheet_by_name(s)
            for j in range(1, sheet.max_column + 1):
                columns.append(get_column_letter(j))
                selecty += '<option value="' + str(j) + '">Column' + str(j) + '</option>'
                selectx += '<option value="' + str(j) + '">Column' + str(j) + '</option>'

        return render(request, 'analysis_index.html', locals())


    elif(request.POST.get('plot') and 'user_name' in request.COOKIES) :
        print('post plot')

        username = 'not a valid user'
        if 'user_name' in request.COOKIES:
            username = request.COOKIES['user_name']
        logintime = '0'
        if 'login_time' in request.COOKIES:
            logintime = request.COOKIES.get('login_time')

        analysistype = request.POST.get('analysistype')

        name=request.POST.get("name")
        xlable=request.POST.get("xlable")
        ylable=request.POST.get("ylable")
        countcols=request.POST.get('countcols')
        print(countcols)
        fig = plt.figure()
        fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold',color='red')
        ax = fig.add_subplot(111)

        fig.subplots_adjust(top=0.85)
        ax.set_xlabel(request.POST.get('xlable'))
        ax.set_ylabel(request.POST.get('ylable'))
        ax.plot()
        plt.grid(True)

        maxx=0
        maxy=0
        minx=0
        miny=0
        ddata=[]

        TOOLS = 'box_select,resize,reset,pan,tap,wheel_zoom,save,hover'

        bplot=figure(title=request.POST.get('name'),plot_width=640,plot_height=480,x_axis_label=request.POST.get('xlable'),y_axis_label=request.POST.get('ylable'),tools=TOOLS, toolbar_location='above')
        bplot.title.text_color="red"
        bplot.title.text_font_size='22px'



        ycolumns=[]
        xcolumns=[]
        colorscol=[]
        x=[]
        y=[]
        for i in countcols:
            ycolumns.append(request.POST.get('ydata'+str(i)))
            colorscol.append(request.POST.get('color'+str(i)))

        dest_filename = 'analysis/static/xlsx/' + username + logintime + '.xlsx'
        wb = openpyxl.load_workbook(dest_filename)
        sheets = wb.get_sheet_names()

        for s in sheets:
            sheet=wb.get_sheet_by_name(s)
            for i in range(1,sheet.max_row+1):
                dtemp=[]
                for j in range(1,sheet.max_column+1):
                    if analysistype!='bar' and analysistype!='pie':
                        dtemp.append(float(sheet.cell(row=i,column=j).value))
                    else:
                        dtemp.append((sheet.cell(row=i, column=j).value))
                ddata.append(dtemp)

        #print(ddata)
        #print('hellooo')
        ddata=str(ddata)
        ddata=ddata[1:-1]


        analysistype = request.POST.get('analysistype')
        print(analysistype)
        ''' plotting data are saved in x and y[]'''
        div_script=''
        div_tag=''
        xcol=request.POST.get('xdata')
        for i in range(1,sheet.max_row+1):
                x.append(float(sheet.cell(row=i,column=int(xcol)).value))



        '''get selected columns data'''

        print(analysistype)
        if analysistype=='linearregression':
            print('linearregressioN')
            xcol = request.POST.get('xdata')
            x=[]
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=int(xcol)).value))
            ycol = request.POST.get('ydata')
            y=[]
            for i in range(1, sheet.max_row + 1):
                y.append(float(sheet.cell(row=i, column=int(ycol)).value))
            y=np.array(y)
            x=np.array(x)
            f=polyfit(x,y,1)
            print(f)
            p = np.poly1d(f)
            yhat = p(x)  # or [p(z) for z in x]
            ybar = np.sum(y) / len(y)  # or sum(y)/len(y)
            ssreg = np.sum((yhat - ybar) ** 2)  # or sum([ (yihat - ybar)**2 for yihat in yhat])
            sstot = np.sum((y - ybar) ** 2)  # or sum([ (yi - ybar)**2 for yi in y])
            rval= ssreg / sstot
            yn=polyval(f,x)
            plt.scatter(x, y, marker='.', color='blue', label="Real")
            plt.plot(x, yn, marker='.', markersize=2, alpha=1, color='red', label='R^2='+str(rval)+str(f))








        elif analysistype=='linearregression':
            print('linearregressioN')
            xcol = request.POST.get('xdata')
            x=[]
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=int(xcol)).value))
            ycol = request.POST.get('ydata')
            y=[]
            for i in range(1, sheet.max_row + 1):
                y.append(float(sheet.cell(row=i, column=int(ycol)).value))
            y=np.array(y)
            x=np.array(x)

            regmodel = LinearRegression(n_jobs=-1)
            regmodel.fit([x], [y])
            ynew = regmodel.predict([x])
            yp=[]
            for p in ynew:
                yp=p
            plt.scatter(x, yp, marker='.', color='blue', label="Predicted Fun")
            plt.plot(x, y, marker='.', markersize=2, alpha=1, color='red', label="Fun")
            dic = {ylable: y, xlable: x}
            df = pd.DataFrame(dic)
            scatr = Scatter(df, x=xlable, y=ylable, color=colorscol[0])

            bplot.line(x, yp, color='red', legend='Predicted Fun' )

            div_script, div_tag = autoload_static(scatr, CDN, "js")
        elif analysistype=='corelation':
            print('corelation')
            x=[]
            for i in range(1, sheet.max_row + 1):
                    x.append(float(sheet.cell(row=i, column=int(xcol)).value))
            y=[]
            ycol = request.POST.get('ydata')
            for i in range(1,sheet.max_row+1):
                    y.append(float(sheet.cell(row=i, column=int(ycol)).value))

            coff=pearsonr(x,y)
            plt.scatter(x, y, color=colorscol[0],label='Corelation:'+str(coff[0]))

            dic = {ylable: y, xlable: x}
            df = pd.DataFrame(dic)
            scatr = Scatter(df, x=xlable, y=ylable, color='black')
            div_script, div_tag = autoload_static(scatr, CDN, "js")

        elif analysistype=='interpolation':
            print('interpolation')
            x=[]
            for i in range(1, sheet.max_row + 1):
                    x.append(float(sheet.cell(row=i, column=int(xcol)).value))
            y=[]
            ycol=request.POST.get('ydata')
            for j in range(1,sheet.max_row+1):
                y.append(float(sheet.cell(row=j, column=int(ycol)).value))

            print(x,y)

            f = interp1d(x, y)
            f2 = interp1d(x, y, kind='cubic')
            minx=min(x)
            maxx=max(x)
            xnew = np.linspace(minx, maxx, num=1000, endpoint=True)
            plt.plot(x,y,'o')
            plt.plot(xnew,f(xnew),'-')
            plt.plot(xnew,f2(xnew),'--')

            plt.scatter(x,y)

            dic = {ylable: y, xlable: x}
            df = pd.DataFrame(dic)
            scatr = Scatter(df, x=xlable, y=ylable, color='black')
            div_script, div_tag = autoload_static(scatr, CDN, "js")



        elif analysistype == 'interpolation2d':
            print('interpolation2d')
            x = []
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=int(xcol)).value))
            y = []
            ycol = request.POST.get('ydata')
            for j in range(1, sheet.max_row + 1):
                y.append(float(sheet.cell(row=j, column=int(ycol)).value))

            z=[]
            zcol = request.POST.get('zdata')
            for j in range(1, sheet.max_row + 1):
                z.append(float(sheet.cell(row=j, column=int(zcol)).value))

            print(x, y,z)

            f = interp2d(x, y, z, kind='cubic')
            minx = min(x)
            maxx = max(x)
            xnew = np.linspace(minx, maxx, num=1000, endpoint=True)
            plt.plot(x, y, 'o')
            plt.plot(xnew, f(xnew), '-')

            plt.scatter(x, y)

            dic = {ylable: y, xlable: x}
            df = pd.DataFrame(dic)
            scatr = Scatter(df, x=xlable, y=ylable, color='black')
            div_script, div_tag = autoload_static(scatr, CDN, "js")




        elif analysistype=='polynomialregression':
            print('polynomialregression')
            xcol = request.POST.get('xdata')
            x = []
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=int(xcol)).value))
            ycol = request.POST.get('ydata')
            y = []
            for i in range(1, sheet.max_row + 1):
                y.append(float(sheet.cell(row=i, column=int(ycol)).value))
            y = np.array(y)
            x = np.array(x)
            #deg = request.POST.get('degree')
            deg=request.POST.get('degree')
            f = polyfit(x, y, deg=int(deg))
            print(f)
            yhat = polyval(f,x)  # or [p(z) for z in x]
            ybar = np.sum(y) / len(y)  # or sum(y)/len(y)
            ssreg = np.sum((yhat - ybar) ** 2)  # or sum([ (yihat - ybar)**2 for yihat in yhat])
            sstot = np.sum((y - ybar) ** 2)  # or sum([ (yi - ybar)**2 for yi in y])
            rval = ssreg / sstot
            plt.scatter(x, y, marker='.', color='blue', label="Real")
            plt.plot(x, yhat, marker='.', markersize=2, alpha=1, color='red', label='R^2=' + str(rval) + str(f))


        elif analysistype=='frequencyhistogram':
            print('frequencyhistogram')

            plt.hist(x,color='#2bb3ff')
            print(x)

            dic = { xlable: x}
            df = pd.DataFrame(dic)
            print(df)
            hist= Histogram(df, values=xlable, color='blue')
            div_script, div_tag = autoload_static(hist, CDN, "js")
        elif analysistype=='cluster':
            print('clusters')
            x=[]
            xcol = request.POST.get('xdata')
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=int(xcol)).value))
            X=[]
            cluster_count=int(request.POST.get('clusters'))
            for j in range(0, len(y)):
                for i in range(0,len(x)):
                    X.append([x[i], y[j][i]])

            X=np.array(X,np.int64)
            print(X)
            y_pred = KMeans(n_clusters=cluster_count, random_state=0).fit_predict(X)

            plt.scatter(X[:, 0], X[:, 1], c=y_pred)
            for j in range(0, len(y)):
                dic = {ylable: y[j], xlable: x}
                df = pd.DataFrame(dic)
                scatr = Scatter(df, x=xlable, y=ylable, color=colorscol[j])
                div_script, div_tag = autoload_static(scatr, CDN, "js")

            af = AffinityPropagation(preference=-5000).fit(X)
            cluster_centers_indices = af.cluster_centers_indices_
            labels = af.labels_
            n_clusters_ = len(cluster_centers_indices)
            plt.figure(1)
            plt.clf()
            colors = cycle('bgrcmykbgrcmykbgrcmykbgrcmyk')
            af = AffinityPropagation(preference=-5000).fit(X)
            cluster_centers_indices = af.cluster_centers_indices_
            labels = af.labels_
            n_clusters_ = len(cluster_centers_indices)
            plt.figure(1)

            plt.clf()
            colors = cycle('bgrcmykbgrcmykbgrcmykbgrcmyk')
            for k, col in zip(range(n_clusters_), colors):
                class_members = labels == k
                cluster_center = X[cluster_centers_indices[k]]
                plt.plot(X[class_members, 0], X[class_members, 1], col + '.')
                plt.plot(cluster_center[0], cluster_center[1], 'o', markerfacecolor=col,
                         markeredgecolor='k', markersize=14)
                for x in X[class_members]:
                    plt.plot([cluster_center[0], x[0]], [cluster_center[1], x[1]], col)
            ax = fig.add_subplot(111)
            ax.set_title('estimated no. of clusters '+str(n_clusters_))
            for k, col in zip(range(n_clusters_), colors):
                class_members = labels == k
                cluster_center = X[cluster_centers_indices[k]]
                bplot.line(X[class_members, 0], X[class_members, 1])
                bplot.line(cluster_center[0], cluster_center[1])
                for x in X[class_members]:
                    bplot.line([cluster_center[0], x[0]], [cluster_center[1], x[1]])
            div_script, div_tag = autoload_static(bplot, CDN, "js")


        else:
            print('else')
            for j in range(0,len(y)) :
                plt.plot(x,y[j],marker='.', markersize=2, alpha=1, color=colorscol[j], label="Fun" + str(j))
                bplot.line(x,y[j],color=colorscol[j],legend='Fun'+str(j))
                if(i==sheet.max_row-1):
                    plt.annotate('fun-'+str(j-1), xy=(float(sheet.cell(row=i, column=1).value), float(sheet.cell(row=i, column=j).value)),xytext=(float(sheet.cell(row=i, column=1).value-(maxx/6)), float(sheet.cell(row=i, column=j).value)),
                            arrowprops=dict(facecolor='black',arrowstyle='->'),horizontalalignment='left',verticalalignment='bottom'
                                         )

        plt.legend(loc=0)
        fig1 = plt.gcf()

        if(FileSystemStorage('/static/img/').exists(username+logintime+'.png')):
            os.remove('static/img/'+username+logintime+'.png')

        fig1.savefig('analysis/static/img/'+username+logintime+'.png', dpi=200)
        fig1.savefig('analysis/static/img/' + username + logintime + '.jpg', dpi=200)
        fig1.savefig('analysis/static/img/' + username + logintime + '.eps', dpi=200)
        fig1.savefig('analysis/static/img/' + username + logintime + '.pdf', dpi=200)
        fig1.savefig('analysis/static/img/' + username + logintime + '.svg', dpi=200)
        fig1.savefig('analysis/static/img/' + username + logintime + '.tiff', dpi=200)
        fig1.savefig('analysis/static/img/' + username + logintime + '.ps', dpi=200)
        fig1.savefig('analysis/static/img/' + username + logintime + '.svgz', dpi=200)
        # fig1.savefig('analysis/static/img/test.pgf', dpi=200)#latex not installed sudo apt-get install texlive-xetex
        fig1.savefig('analysis/static/img/' + username + logintime + '.jpeg', dpi=200)
        # fig1.savefig('analysis/static/img/test.raw', dpi=200)#if needed
        fig1.savefig('analysis/static/img/' + username + logintime + '.tif', dpi=200)

        saved = True

        dis = 'analysis/static/img/' + username+logintime + '.png'
        #shutil.copyfile('analysis/static/img/'+username+logintime+'.png', dis)
        saved = True
        imagelink = 'img/' +username+logintime  + '.png'
        filelink = 'xlsx/'+username+logintime+'.xlsx'

        bplot.legend.location = "top_left"
        bplot.legend.background_fill_color = "white"
        bplot.legend.background_fill_alpha = 0.5


        image = 'static/img/' + username + logintime
        plotanalysis=1
        tabled=1
        #print(div_tag,div_script)
        username='InvalidUser'
        logintime='0'
        if 'user_name' in request.COOKIES:
                username = request.COOKIES['user_name']
        if 'login_time' in request.COOKIES:
                logintime=request.COOKIES['login_time']

        dest_filename = 'analysis/static/xlsx/' + username + logintime + '.xlsx'
        wb = openpyxl.load_workbook(dest_filename)
        sheets = wb.get_sheet_names()

        selectx = ''
        selecty = ''
        columns = []
        for s in sheets:
            sheet=wb.get_sheet_by_name(s)
            for j in range(1, sheet.max_column + 1):
                columns.append(get_column_letter(j))
                selecty += '<option value="' + str(j) + '">Column' + str(j) + '</option>'
                selectx += '<option value="' + str(j) + '">Column' + str(j) + '</option>'

        print(columns)
        print(selectx)
        print(selecty)

        #print(div_tag)
        #print(div_script)
        return render(request,'analysis_index.html', locals())
    else:
        print('nothing')
        ddata = 0
        tabled = 0
        plotanalysis = 0
        username='InvalidUser'
        return redirect('./../?usr=0','analysis_index.html',request)


def interpolation(request):
    if request.method != 'POST':
        return render(request, 'interpolation.html', locals())
    else:
        msg = ""
        msg = "post"
        msg += request.POST.get("name")

        fs = FileSystemStorage(location='analysis/static/xlsx/')
        fs._save("interpolation.xlsx", request.FILES.get("file"))

        fig = plt.figure()
        fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold', color='red')
        ax = fig.add_subplot(111)

        fig.subplots_adjust(top=0.85)
        ax.set_title(request.POST.get('axis'))
        ax.set_xlabel(request.POST.get('xlable'))
        ax.set_ylabel(request.POST.get('ylable'))
        ax.plot()
        plt.grid(True)

        wb = openpyxl.load_workbook(request.FILES.get("file"))
        sheets = wb.get_sheet_names()

        col = []
        colr = ['r', 'b', 'g', 'y']

        for s in sheets:
            sheet = wb.get_sheet_by_name(s)
            msg += str(sheet.max_column)
            maxx = 0
            minx = 999999999
            x = []
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=1).value))
                if (x[i - 1] > maxx):
                    maxx = x[i - 1]
                if (x[i - 1] < minx):
                    minx = x[i - 1]

            y = []
            for j in range(2, sheet.max_column + 1):
                y_tmp = []
                for i in range(1, sheet.max_row + 1):
                    y_tmp.append(float(sheet.cell(row=i, column=j).value))
                y.append(y_tmp)

            for j in range(0, len(y)):
                f = interp1d(x, y[j])
                f2 = interp1d(x, y[j], kind='cubic')
                xnew = np.linspace(minx, maxx, num=1000, endpoint=True)
                plt.plot(x, y[j], 'o', xnew, f(xnew), '-', xnew, f2(xnew), '--')

                msg += 'sinx'
        plt.legend(loc=0)
        fig1 = plt.gcf()

        if (FileSystemStorage('/static/img/').exists('interpolation.png')):
            os.remove('static/img/interpolation.png')

        fig1.savefig('analysis/static/img/interpolation.png', dpi=100)
        saved = True

        msg += "post and saved"

        return render(request, 'interpolation.html', locals())














#clustering
def clustering(request):
    if request.method != 'POST':
        return render(request, 'clustering.html', locals())
    else:
        import numpy as np
        import matplotlib.pyplot as plt

        from sklearn.cluster import KMeans
        from sklearn.cluster import AffinityPropagation
        from sklearn import metrics
        from sklearn.datasets.samples_generator import make_blobs
        #generate sample data

        centers = [[1, 1], [-1, -1], [1, -1]]
        X, labels_true = make_blobs(n_samples=300, centers=centers, cluster_std=0.5,
                                    random_state=0)
        #computer affinity propogation
        af = AffinityPropagation(preference=-50).fit(X)
        cluster_centers_indices = af.cluster_centers_indices_
        labels = af.labels_

        n_clusters_ = len(cluster_centers_indices)

        print('Estimated number of clusters: %d' % n_clusters_)
        print("Homogeneity: %0.3f" % metrics.homogeneity_score(labels_true, labels))
        print("Completeness: %0.3f" % metrics.completeness_score(labels_true, labels))
        print("V-measure: %0.3f" % metrics.v_measure_score(labels_true, labels))
        print("Adjusted Rand Index: %0.3f"
              % metrics.adjusted_rand_score(labels_true, labels))
        print("Adjusted Mutual Information: %0.3f"
              % metrics.adjusted_mutual_info_score(labels_true, labels))
        print("Silhouette Coefficient: %0.3f"
              % metrics.silhouette_score(X, labels, metric='sqeuclidean'))
        #plot the result
        import matplotlib.pyplot as plt
        from itertools import cycle

        plt.close('all')
        plt.figure(1)
        plt.clf()

        colors = cycle('bgrcmykbgrcmykbgrcmykbgrcmyk')
        for k, col in zip(range(n_clusters_), colors):
            class_members = labels == k
            cluster_center = X[cluster_centers_indices[k]]
            plt.plot(X[class_members, 0], X[class_members, 1], col + '.')
            plt.plot(cluster_center[0], cluster_center[1], 'o', markerfacecolor=col,
                     markeredgecolor='k', markersize=14)
            for x in X[class_members]:
                plt.plot([cluster_center[0], x[0]], [cluster_center[1], x[1]], col)

        plt.title('Estimated number of clusters: %d' % n_clusters_)

        plt.legend(loc=0)
        fig1 = plt.gcf()

        if (FileSystemStorage('/static/img/').exists('cluster.png')):
            os.remove('static/img/cluster.png')

        fig1.savefig('analysis/static/img/cluster.png', dpi=200)
        saved = True

        return render(request, 'clustering.html', locals())


def curveFit(request):
    def fitfun(x,b):
        return  np.exp(-b*x)
    def fitfun1(x,a,b):
        return x*a+b
    def fitfun2(x,a,b,c):
        return a*x*x+b*x+c
    def fitfun3(x,a,b,c,d):
        return a*x*x*x+b*x*x+c*x+d
    def fitfun4(x,a,b,c):
        return a*np.cos(2*np.pi*x+b)+c


    if request.method!='POST':
        return render(request, 'curvefitting.html', locals())
    else:
        msg = ""
        msg = "post"
        msg += request.POST.get("name")
        '''
        #fake data createing
        x=np.linspace(0,4,50)
        temp=fitfun(x,2.5,1.3,0.5)
        y=temp+0.25*np.random.normal(size=len(temp))
        #scipy optimise module contains a least square curve fit rutine
        '''


        fs = FileSystemStorage(location='analysis/static/xlsx/')
        fs._save("curve_fit.xlsx", request.FILES.get("file"))

        fig = plt.figure()
        fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold', color='red')
        ax = fig.add_subplot(111)

        fig.subplots_adjust(top=0.85)
        ax.set_title(request.POST.get('axis'))
        ax.set_xlabel(request.POST.get('xlable'))
        ax.set_ylabel(request.POST.get('ylable'))
        ax.plot()
        plt.grid(True)

        wb = openpyxl.load_workbook(request.FILES.get("file"))
        sheets = wb.get_sheet_names()

        col = []
        colr = ['r', 'b', 'g', 'y', 'c', 'm', 'k']
        maxx = 0
        maxy = 0
        minx = 0
        miny = 0

        y = []
        x = []
        for s in sheets:
            sheet = wb.get_sheet_by_name(s)
            msg += str(sheet.max_column)

            # getting x values
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=1).value))

            # getting y values:
            x = np.array(x)

            for j in range(2, sheet.max_column + 1):
                y_temp = []
                for i in range(1, sheet.max_row + 1):
                    y_temp.append(float(sheet.cell(row=i, column=j).value))
                y.append(y_temp)

            print(y)
            for y_val in y:
                y_val = np.array(y_val)
                if (int(request.POST.get('variable')) == 2):
                    fitParams, fitCovariances = curve_fit(fitfun1, x, y_val)
                    # msg += str(fitfun1(x, fitParams[0], fitParams[1]))
                    fit = fitfun1(x, fitParams[0], fitParams[1])
                    line1 = plt.plot(x, fitfun1(x, fitParams[0], fitParams[1]), label="Line 1")
                    err1 = plt.errorbar(x, y_val, fmt='bs', yerr=0.1, label="Error-1")
                elif (int(request.POST.get('variable')) == 3):
                    fitParams, fitCovariances = curve_fit(fitfun2, x, y_val)
                    # msg += str(fitfun2(x, fitParams[0], fitParams[1], fitParams[2]))
                    line2 = plt.plot(x, fitfun2(x, fitParams[0], fitParams[1], fitParams[2]), label="Line 2")
                    err2 = plt.errorbar(x, y_val, fmt='g-.', yerr=0.1, label="Error-2")
                elif (int(request.POST.get('variable')) == 4):
                    fitParams, fitCovariances = curve_fit(fitfun3, x, y_val)
                    # msg += str(fitfun3(x, fitParams[0], fitParams[1], fitParams[2],fitParams[3]))
                    line3 = plt.plot(x, fitfun3(x, fitParams[0], fitParams[1], fitParams[2], fitParams[3]),
                                     label="Line 3")
                    err3 = plt.errorbar(x, y_val, fmt='b^', yerr=0.1, label="Error-3")
                elif ((int(request.POST.get('variable')) == 5)):
                    fitParams, fitCovariances = curve_fit(fitfun4, x, y_val)
                    # msg += str(fitfun3(x, fitParams[0], fitParams[1], fitParams[2],fitParams[3]))
                    err4 = plt.errorbar(x, y_val, fmt='bo', yerr=2.0, label="Error-4")

                    line4 = plt.plot(x, fitfun4(x, fitParams[0], fitParams[1], fitParams[2]),
                                     label="Line 4", color='r', marker='^', markersize=3.0)


                    # plt.ylabel('Temperature (C)', fontsize=16)
                    # plt.xlabel('Time', fontsize=16)
                    # plt.xlim(0, 4.1)

                    # plot data as red square with vertical errors
                    # plot best fit analysis

            fig1 = plt.gcf()
            plt.legend()  # to plot top side square for plot detail
            if (FileSystemStorage('/static/img/').exists('curve_fit.png')):
                os.remove('static/img/curve_fit.png')

            fig1.savefig('analysis/static/img/curve_fit.png', dpi=200)
            saved = True
            return render(request, 'curvefitting.html', locals())



def lineFit(request):
    def fitfun(x,b):
        return  np.exp(-b*x)
    def fitfun1(x,a,b):
        return x*a+b
    def fitfun2(x,a,b,c):
        return a*x*x+b*x+c
    def fitfun3(x,a,b,c,d):
        return a*x*x*x+b*x*x+c*x+d
    def fitfun4(x,a,b,c):
        return a*np.cos(2*np.pi*x+b)+c


    if request.method!='POST':
        return render(request, 'linefitting.html', locals())
    else:
        msg = ""
        msg = "post"
        msg += request.POST.get("name")

        '''
        #fake data createing
        x=np.linspace(0,4,50)
        temp=fitfun(x,2.5,1.3,0.5)
        y=temp+0.25*np.random.normal(size=len(temp))
        #scipy optimise module contains a least square curve fit rutine
        '''


        fs = FileSystemStorage(location='analysis/static/xlsx/')
        fs._save("line_fit.xlsx", request.FILES.get("file"))

        fig = plt.figure()
        fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold', color='red')
        ax = fig.add_subplot(111)

        fig.subplots_adjust(top=0.85)
        ax.set_title(request.POST.get('axis'))
        ax.set_xlabel(request.POST.get('xlable'))
        ax.set_ylabel(request.POST.get('ylable'))
        ax.plot()
        plt.grid(True)

        wb = openpyxl.load_workbook(request.FILES.get("file"))
        sheets = wb.get_sheet_names()

        col = []
        colr = ['r', 'b', 'g', 'y', 'c', 'm', 'k']
        maxx = 0
        maxy = 0
        minx = 0
        miny = 0

        y = []
        x = []
        for s in sheets:
            sheet = wb.get_sheet_by_name(s)
            msg += str(sheet.max_column)

            # getting x values
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=1).value))

            # getting y values:
            x = np.array(x)

            for j in range(2, sheet.max_column + 1):
                y_temp = []
                for i in range(1, sheet.max_row + 1):
                    y_temp.append(float(sheet.cell(row=i, column=j).value))
                y.append(y_temp)

            print(y)
            for y_val in y:
                y_val = np.array(y_val)
                if (int(request.POST.get('variable')) == 2):
                    # parameters
                    a = 0.8
                    b = -4
                    x_val = polyval([a, b], y_val)

                    # Linear regressison -polyfit - polyfit can be used other orders polys
                    (ar, br) = polyfit(x, x_val, 1)
                    x_reg = np.polyval([ar, br], x)
                    # compute the mean square error
                    err = sqrt(sum((x_reg - x_val) ** 2) / i)

                    # matplotlib ploting
                    plt.title('Linear Regression Example')
                    plt.plot(x, x_val, 'g.--')
                    plt.plot(x, x_reg, 'r.-')

                    # plt.show()

                    # Linear regression using stats.linregress
                    slope, intercept, r_value, p_value, std_err = stats.linregress(x, x_val)
                    y_pre = intercept + slope * x
                    plt.plot(x, y_pre, 'b.-')

                    plt.legend(['original', 'regression', 'regression_using_fn'])

                elif (int(request.POST.get('variable')) == 3):
                    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y_val)
                    y_pre = intercept + slope * x
                    plt.plot(x, y_pre, 'b.-')

                    plt.legend(['original', 'regression', 'regression_using_fn'])

                    err1 = plt.errorbar(x, y_val, fmt='bs', yerr=0.1, label="Error-3")
                else:
                    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y_val)
                    y_pre = intercept + slope * x
                    plt.plot(x, y_pre, 'b.-')

                    plt.legend(['original', 'regression', 'regression_using_fn'])

                    err1 = plt.errorbar(x, y_val, fmt='bs', yerr=0.1, label="Error-n")
                    # plt.xlabel('Time', fontsize=16
                    # plt.xlim(0, 4.1)

                    # plot data as red square with vertical errors
                    # plot best fit analysis

            fig1 = plt.gcf()
            # plt.legend()#to plot top side square for plot detail
            if (FileSystemStorage('/static/img/').exists('line_fit.png')):
                os.remove('static/img/line_fit.png')

            fig1.savefig('analysis/static/img/line_fit.png', dpi=1000)
            saved = True
            return render(request, 'linefitting.html', locals())








def contour(request):
    if request.method!='POST':
        return render(request, 'contour.html', locals())
    else:
        msg = ""
        msg = "post"
        msg += request.POST.get("name")

        # fake data createing
        # x=np.linspace(0,4,50)
        # temp=fitfun(x,2.5,1.3,0.5)
        # y=temp+0.25*np.random.normal(size=len(temp))
        # scipy optimise module contains a least square curve fit rutine
        doc = Document()
        doc.name = request.POST.get("contour.xlsx")

        fs = FileSystemStorage(location='analysis/static/xlsx/')
        fs._save("contour.xlsx", request.FILES.get("file"))
        doc.save()

        fig = plt.figure()
        fig.suptitle(request.POST.get('name'), fontsize=22, fontweight='bold', color='red')
        ax = fig.add_subplot(111)

        fig.subplots_adjust(top=0.85)
        ax.set_title(request.POST.get('axis'))
        ax.set_xlabel(request.POST.get('xlable'))
        ax.set_ylabel(request.POST.get('ylable'))
        ax.plot()
        plt.grid(True)

        wb = openpyxl.load_workbook(request.FILES.get("file"))
        sheets = wb.get_sheet_names()

        col = []
        colr = ['r', 'b', 'g', 'y']

        y = []
        x = []
        for s in sheets:
            sheet = wb.get_sheet_by_name(s)
            msg += str(sheet.max_column)
            # getting x values
            for i in range(1, sheet.max_row + 1):
                x.append(float(sheet.cell(row=i, column=1).value))

            # getting y values:
            for j in range(2, sheet.max_column + 1):
                y_temp = []
                for i in range(1, sheet.max_row + 1):
                    y_temp.append(float(sheet.cell(row=i, column=j).value))
                y.append(y_temp)
            fig_count = []
            if len(y) > 0:
                m = True

            for j in range(0, len(y)):
                X, Y = np.meshgrid(x, y[j])
                Z1 = mlab.bivariate_normal(X, Y, 1.0, 1.0, 0.0, 0.0)
                Z2 = mlab.bivariate_normal(X, Y, 1.5, 0.5, 1, 1)
                # difference of Gaussians
                Z = 10.0 * (Z2 - Z1)

                # Create a simple contour plot with labels using default colors.  The
                # inline argument to clabel will control whether the labels are draw
                # over the line segments of the contour, removing the lines beneath
                # the label
                fig_count.append('/img/contour' + str(j) + '.png')
                CS = plt.contour(X, Y, Z)
                plt.clabel(CS, inline=1, fontsize=10)
                plt.title('Simplest default with labels')
                fig = plt.gcf()
                # plt.legend()#to plot top side square for plot detail
                if (FileSystemStorage('analysis/static/img/').exists('contour' + str(j) + '.png')):
                    os.remove('analysis/static/img/contour' + str(j) + '.png')

                fig.savefig('analysis/static/img/contour' + str(j) + '.png', dpi=200)
                saved = True

        ''' working contour plot

        delta = 0.025
        x = np.arange(-3.0, 3.0, delta)
        y = np.arange(-2.0, 2.0, delta)
        X, Y = np.meshgrid(x, y)
        Z1 = mlab.bivariate_normal(X, Y, 1.0, 1.0, 0.0, 0.0)
        Z2 = mlab.bivariate_normal(X, Y, 1.5, 0.5, 1, 1)
        # difference of Gaussians
        Z = 10.0 * (Z2 - Z1)

        # Create a simple contour plot with labels using default colors.  The
        # inline argument to clabel will control whether the labels are draw
        # over the line segments of the contour, removing the lines beneath
        # the label
        plt.figure()
        CS = plt.contour(X, Y, Z)
        plt.clabel(CS, inline=1, fontsize=10)
        plt.title('Simplest default with labels')
        # make a colorbar for the contour lines
        CB = plt.colorbar(CS, shrink=0.8, extend='both')

        plt.title('Lines with colorbar')
        plt.hot()  # Now change the colormap for the contour lines and colorbar
        #plt.flag()

        #plt.show()
        fig = plt.gcf()
        # plt.legend()#to plot top side square for plot detail
        if (FileSystemStorage('analysis/static/img/').exists('contour'+str(j)+'.png')):
            os.remove('analysis/static/img/contour'+str(j)+'.png')

        fig.savefig('analysis/static/img/contour'+str(j)+'.png', dpi=200)
        saved = True
        '''

        return render(request, 'contour.html', locals())
